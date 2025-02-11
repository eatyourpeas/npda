# import types
from collections import defaultdict
from typing import Any, Dict, List, Union
import io

from openpyxl.worksheet.worksheet import Worksheet

# import models
from ..models.submission import Submission

# import functions
from project.npda.general_functions.csv.csv_parse import csv_parse

# import third-party libaries
import pandas as pd
from json import loads as json_loads
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment

# import csv mappings
from ...constants.csv_headings import (
    CSV_HEADING_OBJECTS,
    UNIQUE_IDENTIFIER_ENGLAND,
    UNIQUE_IDENTIFIER_JERSEY,
)


def write_errors_to_xlsx(
    errors: dict[str, dict[str, list[str]]],
    original_csv_file_bytes: bytes,
    is_jersey: bool,
) -> bytes:
    """
    Write errors to an Excel file. Highlight invalid cells in the source CSV.

    Args:
      errors A nested dictionary containing errors grouped by row index, then field.

    """

    xlsx_file = io.BytesIO()

    # Get original data
    df = csv_parse(
        io.BytesIO(initial_bytes=original_csv_file_bytes), is_jersey=is_jersey
    ).df
    # Write an xlsx of the original data.
    df.to_excel(xlsx_file, sheet_name="Uploaded data (raw)", index=False)

    # If the csv file is from Jersey, add the Jersey unique identifier to the CSV headings, otherwise add the England unique identifier
    df_errors = flatten_errors(
        errors=errors,
        original_data=df,
        identifier_field="Unique Reference Number" if is_jersey else "NHS Number",
        csv_headings=(UNIQUE_IDENTIFIER_JERSEY if is_jersey else UNIQUE_IDENTIFIER_ENGLAND) + CSV_HEADING_OBJECTS
    )

    print(df_errors.to_string())

    # Add sheet that lists the errors.
    with pd.ExcelWriter(xlsx_file, mode="a", engine="openpyxl") as writer:
        df_errors.to_excel(writer, sheet_name="Errors - Overview", index=False)

    # Load the workbook in openpyxl
    wb: Workbook = load_workbook(xlsx_file)

    # Set text to red
    overview_sheet = wb["Errors - Overview"]
    for row in overview_sheet.iter_rows(min_row=2, min_col=4):
        for cell in row:
            cell.font = Font(color="FF0000")

    # Setup the styled worksheet
    styled_sheet: Worksheet = wb.copy_worksheet(wb["Uploaded data (raw)"])
    styled_sheet.title = (
        "Uploaded data (comments)"  # You can set any name for the copied sheet
    )

    # Style the openpyxl worksheet to highlight in red erroneous/invalid cells.
    # Also add comments to annotate the actual error.
    for _, patient_errors in df_errors.iterrows():
        row_index = patient_errors["Original CSV Row"] + 1

        field_name = patient_errors["Column"]
        field_errors = patient_errors["Errors"]

        column_index = find_column_index_by_name(field_name, styled_sheet)
        if column_index:
            styled_sheet.cell(row=row_index, column=column_index).fill = (
                PatternFill(patternType="solid", fgColor="FFC9C9")
            )  # Change color to red.
            styled_sheet.cell(row=row_index, column=column_index).comment = Comment(
                field_errors,
                "Data Validator [Automated: RCPCH]",
                height=300,
                width=300,
            )

    # Specify the desired order by reordering the `workbook.worksheets` list
    wb._sheets = [
        wb["Uploaded data (raw)"],
        wb["Uploaded data (comments)"],
        wb["Errors - Overview"],
    ]

    # Save the styled sheet.
    xlsx_file = io.BytesIO()
    wb.save(xlsx_file)

    return xlsx_file.getvalue()


def find_column_index_by_name(column_name: str, ws: Worksheet) -> int | None:
    column_index = None
    for col in ws.iter_cols(
        1, ws.max_column, 1, 1
    ):  # Check headers in the first row only
        if col[0].value == column_name:
            column_index = col[0].column  # Get the column index
            break
    return column_index


def model_field_to_csv_heading(model_field: str) -> str:
    match model_field:
        case 'nhs_number':
            return 'NHS Number'
        case 'unique_reference_number':
            return 'Unique Reference Number'
        case _:
            return next(
                (item["heading"] for item in CSV_HEADING_OBJECTS if item["model_field"] == model_field),
                model_field
            )


def flatten_errors(
    #  {row_number: {field_name: [error_messages]}}
    errors: dict[int, dict[str, list[str]]],
    original_data: pd.DataFrame,
    identifier_field: str,
    csv_headings: List[Dict[str, str]],
) -> pd.DataFrame:
    identifier_column = model_field_to_csv_heading(identifier_field)

    rows = []

    for row_ix, row_errors in errors.items():
        for field, errors in row_errors.items():
            rows.append({
                "Original CSV Row": int(row_ix) + 1,
                identifier_field: original_data.loc[int(row_ix), identifier_column],
                "Column": model_field_to_csv_heading(field),
                "Errors": "; ".join(errors),
            })

    return pd.DataFrame(rows, columns=[
        "Original CSV Row",
        model_field_to_csv_heading(identifier_field),
        "Column",
        "Errors"
    ])
